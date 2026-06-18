import os
import webbrowser
import json
from datetime import datetime
from time import perf_counter
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

from agents.market_data import MarketDataAgent
from agents.indicator_agent import IndicatorAgent
from agents.pattern_agent import PatternAgent
from agents.scoring_agent import ScoringAgent
from agents.decision_agent import DecisionAgent
from agents.benchmark_agent import BenchmarkAgent
from agents.ranking_agent import RankingAgent
from agents.dashboard_agent import DashboardAgent
from agents.ai_technical_agent import AITechnicalAgent

from utils.relative_strength import calculate_relative_strength
from utils.volume import volume_score
from utils.trend_strength import trend_score
from utils.atr import atr_percent
from utils.risk import risk_rating
from utils.support_resistance import get_support, get_resistance
from utils.breakout_confirmation import confirmed_breakout


def create_folders():

    os.makedirs("reports/excel", exist_ok=True)

    os.makedirs("reports/json", exist_ok=True)

    os.makedirs("reports/html", exist_ok=True)


def main():

    create_folders()

    market = MarketDataAgent()
    indicator_agent = IndicatorAgent()
    pattern_agent = PatternAgent()
    scoring_agent = ScoringAgent()
    decision_agent = DecisionAgent()
    benchmark_agent = BenchmarkAgent()
    ranking_agent = RankingAgent()
    dashboard_agent = DashboardAgent()
    ai_agent = AITechnicalAgent()

    results = []
    errors = []
    stock_data = {}

    symbols = pd.read_csv("data/nifty500.csv")
    benchmark_df = benchmark_agent.fetch_nifty()

    print("\n")
    print("=" * 80)
    print("STOCK PATTERN SCANNER")
    print("=" * 80)

    for idx, symbol in enumerate(symbols["symbol"], start=1):

        try:

            print(f"\n[{idx}/{len(symbols)}] {symbol}")
            print(f"\nScanning {symbol}...")

            # --------------------------------
            # Fetch Market Data
            # --------------------------------

            df = market.fetch(symbol)

            if df.empty:
                print(f"No data found for {symbol}")
                continue

            # --------------------------------
            # Indicators
            # --------------------------------

            df = indicator_agent.enrich(df)
            if df.empty:
                print(f"Insufficient data for {symbol}")
                continue

            latest = df.iloc[-1]

            # --------------------------------
            # Pattern Detection
            # --------------------------------

            patterns = pattern_agent.scan(df)

            # --------------------------------
            # Scoring
            # --------------------------------

            score = scoring_agent.score(patterns, df)

            atr_pct = atr_percent(df)
            risk = risk_rating(atr_pct)
            support = get_support(df)
            resistance = get_resistance(df)
            close_price = float(latest["Close"])
            rs = calculate_relative_strength(df, benchmark_df)
            volume_ratio = volume_score(df)
            trend = trend_score(df)
            breakout_confirmed = (confirmed_breakout(close_price, resistance, volume_ratio, rs))

            final_score = ranking_agent.rank(
                technical_score=score, 
                rs_score=rs, 
                trend_score=trend, 
                volume_ratio=volume_ratio, 
                risk=risk,
                breakout_confirmed= breakout_confirmed)

            if final_score < 60:
                continue

            # --------------------------------
            # Volatility
            # --------------------------------

            atr_pct = atr_percent(df)
            risk = risk_rating(atr_pct)

            # --------------------------------
            # Decision
            # --------------------------------

            decision = (decision_agent.decide(final_score))

            # --------------------------------
            # Pattern Names
            # --------------------------------

            pattern_names = (", ".join(p["pattern"] for p in patterns) if patterns else "None")

            # --------------------------------
            # Store Result
            # --------------------------------

            results.append({
                "Symbol": symbol,
                "Close": round(float(latest["Close"]), 2),
                "RSI": round(float(latest["RSI"]), 2),
                "SMA20": round(float(latest["SMA20"]), 2),
                "SMA50": round(float(latest["SMA50"]), 2),
                "SMA200": round(float(latest["SMA200"]), 2),
                "Patterns": pattern_names,
                "TechnicalScore": score,
                "Decision": decision,
                "RS": rs,
                "VolumeRatio": volume_ratio,
                "TrendScore": trend,
                "FinalScore": final_score,
                "Patterns": pattern_names,
                "Support": support if support else 0,
                "Resistance": resistance if resistance else 0,
                "ATRPercent": atr_pct,
                "Risk": risk,
                "BreakoutConfirmed": breakout_confirmed,
            })

            print(f"FinalScore={final_score}    Decision={decision}")

            stock_data[symbol] = {
                "df": df,
                "support": support,
                "resistance": resistance,
                "rs": rs,
                "atr_pct": atr_pct
            }

        except Exception as ex:
            print(f"Error processing {symbol}: {str(ex)}")
            errors.append({"Symbol": symbol, "Error": str(ex)})

    # ====================================
    # DataFrame
    # ====================================

    results_df = pd.DataFrame(results)

    if results_df.empty:
        print("\nNo results generated.")
        return

    # ====================================
    # Sort by Score
    # ====================================

    results_df = (results_df.sort_values(by="FinalScore", ascending=False).reset_index(drop=True))

    # ====================================
    # Display Top Stocks
    # ====================================

    print("\n")
    print("=" * 80)
    print("TOP STOCKS")
    print("=" * 80)

    # print(results_df[["Symbol", "TechnicalScore", "RS", "TrendScore", "VolumeRatio", "FinalScore", "Decision"]].head(10))
    top10 = results_df.head(10)

    for idx, row in top10.iterrows():

        print(
            f"{idx+1}. "
            f"{row['Symbol']} | "
            f"{row['FinalScore']} | "
            f"{row['Decision']}"
        )

    # ====================================
    # Elite Picks
    # ====================================
    elite = results_df[(results_df["BreakoutConfirmed"] == True) & (results_df["FinalScore"] >= 85)]    

    # ====================================
    # Excel Export
    # ====================================

    timestamp = (datetime.now().strftime("%Y%m%d_%H%M%S"))

    excel_file = (f"reports/excel/scan_{timestamp}.xlsx")

    results_df.to_excel(excel_file, index=False)

    top20 = results_df[results_df["FinalScore"] >= 85].head(20)
    ai_candidates = results_df[(results_df["BreakoutConfirmed"] == True) & (results_df["FinalScore"] >= 90)]
    ai_candidates = ai_candidates.head(10)
    ai_candidates.to_excel(f"reports/excel/ai_candidates_{timestamp}.xlsx", index=False)
    
    top20.to_excel(f"reports/excel/top20_{timestamp}.xlsx", index=False)

    # ====================================
    # AI Analysis
    # ====================================
    ai_start = perf_counter()
    ai_results = []
    for symbol in ai_candidates["Symbol"]:
        try:
            print(f"AI Analyzing {symbol}")

            data = stock_data[symbol]

            response = ai_agent.analyze(
                symbol=symbol,
                df=data["df"],
                support=data["support"],
                resistance=data["resistance"],
                rs=data["rs"],
                atr_pct=data["atr_pct"]
            )

            response = response.strip()
            response = response.replace("```json", "")
            response = response.replace("```", "")
            response = response.strip()

            parsed = json.loads(response)

            ai_results.append({
                "Symbol": symbol,
                "AIPatterns": ",".join(parsed.get("patterns", [])),
                "AIConfidence": parsed.get("confidence", 0),
                "AITrend": parsed.get("trend", ""),
                "AIReasoning": parsed.get("reasoning", "")
            })

        except Exception as ex:
            print(f"AI Error {symbol}: {ex}")
    
    ai_end = perf_counter()
    print(f"\nAI Time: {round(ai_end-ai_start,2)} sec")

    ai_df = pd.DataFrame(ai_results)
    
    ai_json_file = (f"reports/json/ai_analysis_{timestamp}.json")
    ai_df.to_json(ai_json_file, orient="records", indent=4)

    results_df = results_df.merge(ai_df, on="Symbol",how="left")

    elite = results_df[(results_df["BreakoutConfirmed"] == True) & (results_df["FinalScore"] >= 85)]

    elite.to_excel(f"reports/excel/elite_{timestamp}.xlsx", index=False)

    wb = load_workbook(excel_file)
    ws = wb.active

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3"
    )
    bold_font = Font(
        bold=True
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font

    for column in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[column[0].column_letter].width = length + 2

    wb.save(excel_file)

    summary_sheet = (wb.create_sheet("Summary"))
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Total Stocks", len(results_df)])
    summary_sheet.append(["Strong Buy", len(results_df[results_df["Decision"] == "STRONG BUY"])])
    summary_sheet.append(["Buy", len(results_df[results_df["Decision"] == "BUY"])])
    summary_sheet.append(["Watch", len(results_df[results_df["Decision"] == "WATCH"])])
    summary_sheet.append(["Watch", len(results_df[results_df["Decision"] == "AVOID"])])

    wb.save(excel_file)

    # ====================================
    # Top Picks
    # ====================================

    top_picks = (results_df[results_df["FinalScore"] >= 80])

    json_file = (f"reports/json/top_picks_{timestamp}.json")

    top_picks.to_json(json_file, orient="records", indent=4)

    top20.to_json(f"reports/json/top20_{timestamp}.json", orient="records", indent=4)

    # ====================================
    # Daily Summary
    # ====================================

    summary = {

        "Total Stocks": len(results_df),
        "Strong Buy": len(results_df[results_df["Decision"] == "STRONG BUY"]),
        "Buy": len(results_df[results_df["Decision"] == "BUY"]),
        "Watch": len(results_df[results_df["Decision"] == "WATCH"]),
        "Avoid": len(results_df[results_df["Decision"] == "AVOID"])
    }

    print("\n")
    print("=" * 80)
    print("DAILY SUMMARY")
    print("=" * 80)

    for key, value in summary.items():
        print(f"{key}: {value}")

    print("\n")
    print("=" * 80)
    print("SCAN COMPLETED")
    print("=" * 80)

    if errors:
        pd.DataFrame(errors).to_excel("reports/excel/errors.xlsx", index=False)


    # ====================================
    # HTML Dahsboard Generation
    # ====================================
    html_file = (f"reports/html/dashboard_{timestamp}.html")
    dashboard_agent.generate(results_df, elite, html_file)
    webbrowser.open("file://" + os.path.abspath(html_file))


if __name__ == "__main__":
    start_time = perf_counter()

    main()

    end_time = perf_counter()
    print(f"\nScan Time: {round(end_time - start_time,2)} sec")